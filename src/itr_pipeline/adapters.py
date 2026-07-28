from __future__ import annotations

import csv
import json
import os
import re
import urllib.parse
import urllib.request
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any

from .models import LedgerRecord


class SourceAdapter(ABC):
    @abstractmethod
    def load(self) -> list[LedgerRecord]:
        """Read the authoritative ledger without modifying it."""


class JsonLedgerAdapter(SourceAdapter):
    def __init__(self, path: Path):
        self.path = path

    def load(self) -> list[LedgerRecord]:
        payload = json.loads(self.path.read_text(encoding="utf-8"))
        rows = payload["records"] if isinstance(payload, dict) else payload
        return [LedgerRecord.from_dict(row) for row in rows]


class CsvLedgerAdapter(SourceAdapter):
    def __init__(self, path: Path):
        self.path = path

    def load(self) -> list[LedgerRecord]:
        with self.path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [
                LedgerRecord.from_dict(_prepare_tabular_row(row, row_number))
                for row_number, row in enumerate(csv.DictReader(handle), start=2)
                if any(value not in (None, "") for value in row.values())
            ]


class XlsxLedgerAdapter(SourceAdapter):
    def __init__(self, path: Path, sheet: str = ""):
        self.path = path
        self.sheet = sheet

    def load(self) -> list[LedgerRecord]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "Excel import requires openpyxl. Install desktop requirements or use CSV/JSON."
            ) from exc

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            if self.sheet:
                if self.sheet not in workbook.sheetnames:
                    raise ValueError(f"worksheet not found: {self.sheet}")
                worksheet = workbook[self.sheet]
            else:
                worksheet = workbook.active
            iterator = worksheet.iter_rows(values_only=True)
            try:
                headers = [str(value or "").strip() for value in next(iterator)]
            except StopIteration:
                return []
            rows = []
            for row_number, values in enumerate(iterator, start=2):
                if not any(value not in (None, "") for value in values):
                    continue
                rows.append(
                    LedgerRecord.from_dict(
                        _prepare_tabular_row(dict(zip(headers, values)), row_number)
                    )
                )
            return rows
        finally:
            workbook.close()


def _prepare_tabular_row(raw: dict[str, Any], row_number: int) -> dict[str, Any]:
    """Normalize portable CSV/XLSX columns without project-specific assumptions."""
    row = {str(key or "").strip(): value for key, value in raw.items()}
    lowered = {key.casefold(): key for key in row}

    def copy_alias(target: str, *aliases: str) -> None:
        if target in row and row[target] not in (None, ""):
            return
        for alias in aliases:
            source = lowered.get(alias.casefold())
            if source is not None and row[source] not in (None, ""):
                row[target] = row[source]
                return

    copy_alias("task_id", "task id", "row id")
    copy_alias("itr_id", "itr id", "ir", "ir no", "ir number", "inspection request")
    copy_alias("submitted_date", "submitted date", "submission date", "date")
    copy_alias("site", "location")
    copy_alias("process", "activity")
    copy_alias("chainage", "mileage")
    row.setdefault("task_id", f"ROW-{row_number:05d}")
    row.setdefault("site", "Imported")
    row.setdefault("process", "Inspection")

    history = row.get("revision_history")
    if isinstance(history, str):
        history = json.loads(history or "{}")
    if not isinstance(history, dict):
        history = {}
    for key, value in row.items():
        normalized = str(key).strip().upper().replace("_", "-").replace(" ", "-")
        if re.fullmatch(r"REV-\d{2}", normalized):
            history[normalized] = value
    if not history:
        history = {"REV-00": row.get("status", ""), "REV-01": ""}
    row["revision_history"] = history
    return row


class PortalAdapter(ABC):
    @abstractmethod
    def preflight(self) -> None:
        """Validate configuration before any query."""

    @abstractmethod
    def query(self, record: LedgerRecord) -> dict[str, Any]:
        """Return match_code and submission history for one record."""


class FixturePortalAdapter(PortalAdapter):
    def __init__(self, path: Path):
        self.path = path
        self._rows: dict[str, dict[str, Any]] = {}

    def preflight(self) -> None:
        payload = json.loads(self.path.read_text(encoding="utf-8"))
        if isinstance(payload, dict) and "results" not in payload:
            self._rows = {str(key): value for key, value in payload.items()}
        else:
            rows = payload["results"] if isinstance(payload, dict) else payload
            self._rows = {str(row["task_id"]): row for row in rows}

    def query(self, record: LedgerRecord) -> dict[str, Any]:
        if record.task_id not in self._rows:
            raise LookupError(f"portal fixture has no result for task_id={record.task_id}")
        return self._rows[record.task_id]


class HttpJsonPortalAdapter(PortalAdapter):
    """Adapter for a private, sanitized JSON gateway."""

    def __init__(self, base_url: str, token_env: str = ""):
        self.base_url = base_url
        self.token_env = token_env

    def preflight(self) -> None:
        parsed = urllib.parse.urlparse(self.base_url)
        if parsed.scheme not in {"http", "https"} or not parsed.netloc:
            raise ValueError("crawler.base_url must be an absolute HTTP(S) URL")
        if self.token_env and not os.getenv(self.token_env):
            raise ValueError(f"required environment variable is missing: {self.token_env}")

    def query(self, record: LedgerRecord) -> dict[str, Any]:
        url = f"{self.base_url}?{urllib.parse.urlencode({'itr_id': record.itr_id})}"
        headers = {"Accept": "application/json"}
        if self.token_env:
            headers["Authorization"] = f"Bearer {os.environ[self.token_env]}"
        request = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))


def build_source_adapter(root: Path, config: dict[str, Any]) -> SourceAdapter:
    path = root / config["path"]
    adapter = config["adapter"]
    if adapter == "json-ledger":
        return JsonLedgerAdapter(path)
    if adapter == "csv-ledger":
        return CsvLedgerAdapter(path)
    if adapter == "xlsx-ledger":
        return XlsxLedgerAdapter(path, str(config.get("sheet", "")))
    raise ValueError(f"unsupported source adapter: {adapter}")


def build_portal_adapter(root: Path, config: dict[str, Any]) -> PortalAdapter:
    if config["adapter"] == "fixture":
        return FixturePortalAdapter(root / config["path"])
    if config["adapter"] == "http-json":
        return HttpJsonPortalAdapter(str(config["base_url"]), str(config.get("token_env", "")))
    raise ValueError(f"unsupported crawler adapter: {config['adapter']}")
